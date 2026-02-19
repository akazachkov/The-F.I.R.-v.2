import os
import re
import shutil
import tkinter as tk
import openpyxl
import warnings  # Подавление предупреждений, при работе с файлами .xlsb
from pyxlsb import open_workbook
from datetime import datetime
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl import load_workbook, Workbook
from tkinter import messagebox, filedialog
from pathlib import Path
from glob import glob  # Для пакетной обработки папок УС

"""
The F.I.R. (File Interaction Runner) v1.4.0 (16.06.2025)

Архивная версия приложения. Изначально писалась в спешке, без попыток следовать
хоть каким-то "мерам приличия" и "хорошему тону", монолитом, с натыканой во
многих местах внутренней рабочей информацией. Всё это привело к тому, что при
"стерилизации", приложение перестало работать.
Сохранено только ради возможности вспомнить "как что было реализовано" и
возможно, забрать какие-то блоки кода или задумки, в "P.R.I.S.M."

Для работы в не-Win системе, добавить импорт и разблокировать строки в
функциях.
import subprocess  # Модуль используется для открытия файла в зависимости
от операционной системы
import sys  # Модуль для проверки типа операционной системы
"""


# Функция для кнопки 1 - 'Обработать файлы в директории'. Работа с file_1
def copy_file_1_button(content_fwt_rrp, content_fwt_2025):
    suffix = '-file_1_customer'
    # Получаем путь из поля ввода
    source_directory_file_1 = entry.get()

    # Получаем список файлов в исходной директории
    files = [
        f for f in os.listdir(source_directory_file_1) if os.path.isfile(
            os.path.join(source_directory_file_1, f))]

    # Если файлов нет - выводим сообщение и завершаем функцию
    if not files:
        print('Файлов не обнаружено')
        return

    # Проверка наличия лог-файла и создание, если его нет
    log_path = os.path.join(source_directory_file_1, '_log_file_1.xlsx')
    if not os.path.exists(log_path):
        create_log_file(log_path)

    # Открытие лог-файла для добавления записей
    log_wb = load_workbook(log_path)

    # Получаем базовую (родительскую) директорию
    base_directory: str = os.path.dirname(source_directory_file_1)

    # Загружаем таблицу для определения организации
    wb = load_workbook(content_fwt_2025)
    ws = wb['customer']
    org_mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        current_id = row[0]  # Столбец - ID
        org = row[46] if len(row) > 47 else ''  # Столбец - Организация
        org_mapping[str(current_id)] = org if org in [
            'contractor_1', 'contractor_2'] else 'Не определено'
    wb.close()

    for file in files:
        # Определяем имя исходного файла, включая путь до него
        input_filename = os.path.join(source_directory_file_1, file)

        # Получаем имя файла без расширения
        file_name, extension = os.path.splitext(file)

        # Пропускаем лог-файл и работаем с остальными файлами в папке
        if file == '_log_file_1.xlsx':
            continue

        # Извлекаем часть до первого тире (если есть)
        if '-' in file_name:
            base_name = file_name.split('-')[0].strip()
        else:
            base_name = file_name

        # Дополняем base_name нулями до 4 символов
        file_name = base_name.zfill(4)

        # Определяем организацию из таблицы (удаляем ведущие нули)
        org = org_mapping.get(file_name.lstrip('0'), 'Не определено')

        # Создаем поддиректорию "Для загрузки" в исходной директории
        upload_dir = os.path.join(
            source_directory_file_1, 'Для загрузки', org)
        os.makedirs(upload_dir, exist_ok=True)

        # Копируем файл в поддиректорию "Для загрузки"
        upload_path = os.path.join(upload_dir, file)
        shutil.copy2(input_filename, upload_path)

        # Добавляем суффикс перед расширением
        new_file_name = f'{file_name}{suffix}{extension}'

        # Определяем путь к директории назначения
        destination_folder = os.path.join(base_directory, file_name)

        # Создаем директорию назначения, если ее нет
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)

        # Формируем полный путь к файлу в новой директории
        destination = os.path.join(destination_folder, new_file_name)

        # Формируем имя файла для списка на вынос
        output_file_inserted_equipment = os.path.join(
            destination_folder,
            f'{file_name} Список{extension}')
        transform_excel_list(
            input_filename,
            output_file_inserted_equipment)

        # Формируем имя файла file_2
        output_file_removal_equipment = os.path.join(
            destination_folder,
            f'{file_name}-file_2_customer{extension}')
        transform_excel_removal_equipment(
            input_filename,
            output_file_removal_equipment,
            content_fwt_rrp)

        # Перемещаем файл в целевую директорию
        shutil.move(input_filename, destination)

        # Добавляем запись в лог-файл о завершении перемещения
        log_wb.active.append([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Обработан и перемещен\n',
            file_name,
            destination,
            suffix[1:],
            org])

        # Добавляем имя обработанного файла в текстовое поле
        text_area.insert(tk.END, f'{file_name}\n')

    # Добавляем строку в качестве разделителя и закрываем лог-файл
    log_wb.active.append(['работа завершена'])
    log_wb.save(log_path)
    log_wb.close()

    messagebox.showinfo('Результат', 'Дело сделано')


# Функция для кнопки 2 - указание пути к лог-файлу по file_1
def open_log_file_file_1():
    # Получаем путь из поля ввода
    directory_path = entry.get()
    # Формируем путь к лог-файлу
    log_file_path = os.path.join(directory_path, '_log_file_1.xlsx')
    # Передаем путь к директории в функцию
    check_and_open_log(log_file_path)


# Функция для кнопки 3 - открытие окна выбора file_2 и формирование списков
def open_file_file_2():
    input_filename = filedialog.askopenfilename(
        title='Выберите файл',
        filetypes=(('Все файлы', '*.*'), ('Файлы Excel', '*.xlsx')))

    # Проверяем, был ли выбран файл
    if input_filename:

        # Получаем путь к директории и разбиваем имя файла на имя и расширение
        input_dir = os.path.dirname(input_filename)
        input_basename = os.path.basename(input_filename)
        filename, file_extension = os.path.splitext(input_basename)

        # Создаем имя выходного файла, включая путь до него
        output_filename_1 = os.path.join(
            input_dir,
            f'{filename[:4]} Список на вынос{file_extension}')

        # Создание из исходного файла, списка на вынос
        transform_excel_list(input_filename, output_filename_1)

        # Открываем директорию назначения (возможно открытие файла)
        open_file_and_folder(output_filename_1)
    else:
        print('Файл не выбран')


# Функция для кнопки 4 - поиск и копирование 1 (при отсутствии - 2)
def copy_ktd_button(doc24_dir, doc25_dir, dest_dir_1, dest_dir_2):
    """
    Принцип работы с файлами 1 и 2:
    1) проверяем наличие данных в table_path, на листе 1:
    1.1) пропускаем строки, в которых нет данных в столбцах 0 или 57 и 31;
    1.2) если в столбцах 0 и 57 есть ли данные, то проверяем год, в дате из
    столбца 57:
    1.2.1) если год 2024 - ищем файл .pdf в директории doc24_dir / 9. 1:
    1.2.1.1) при нахождении одного файла .pdf, копируем его в папку
    dest_dir_rd и вносим запись в лог-файл с действием "файл скопирован"
    1.2.1.2) если файлов .pdf обнаружено больше одного, то проверяется дата
    изменения файлов и копируется более новый, вносим запись в лог-файл с
    действием "файл скопирован"
    1.2.1.3) если файлы .pdf отсутствуют, вносим запись в лог-файл с действием
    "файл не обнаружен"
    1.2.2) если год 2025 - ищем файл .pdf в директории doc25_dir / 9. 1:
    1.2.2.1) при нахождении одного файла .pdf, копируем в папку dest_dir_1 и
    вносим запись в лог-файл с действием "файл скопирован"
    1.2.2.2) если файлов .pdf обнаружено больше одного, то проверяется дата
    изменения файлов и копируется более новый, вносим запись в лог-файл с
    действием "файл скопирован"
    1.2.3) если год 2025, но файл в директории doc25_dir / 9. 1 отсутствует,
    то ищем файл .pdf в директории doc24_dir / 9. 1 - при нахождении файла
    .pdf, проверяем дату в имени файла. Дата всегда указана в начале имени
    файла (учтены различные варианты написания даты - 01.01.2025, 01-01-2025,
    01.01.25, 01-01-25):
    1.2.3.1) если год соответствует, копируем в папку dest_dir_1 и вносим
    запись в лог-файл с действием "файл скопирован"
    1.2.3.2) если в дате в имени файла указан не 2025, то файл не копируем и
    вносим запись в лог-файл с действием "файл не обнаружен"
    1.2.3.3) если файлов .pdf обнаружено больше одного, то проверяется дата
    изменения файлов и копируется более новый, вносим запись в лог-файл с
    действием "файл скопирован"
    1.2.4) если год 2025, но файл в директориях doc25_dir / 9. 1 и
    doc24_dir / 9. 1 отсутствует - вносим запись в лог-файл с действием
    "файл не обнаружен"
    1.3) если в столбце 0 есть данные, но в столбце 57 нет данных, то
    проверяем наличие даты в столбце 31, в дате из столбца 31 смотрим год:
    1.3.1) если год 2024 - ищем файл .pdf в директории doc24_dir / 7. 2:
    1.3.1.1) при нахождении одного файла .pdf, копируем его в папку
    dest_dir_2 и вносим запись в лог-файл с действием "файл скопирован"
    1.3.1.2) если файлов .pdf обнаружено больше одного, то проверяется дата
    изменения файлов и копируется более новый, вносим запись в лог-файл с
    действием "файл скопирован"
    1.3.1.3) если файлы .pdf отсутствуют, вносим запись в лог-файл с действием
    "файл не обнаружен"
    1.3.2) если год 2025 - ищем файл .pdf в директории doc25_dir / 7. 2:
    1.3.2.1) при нахождении одного файла .pdf, копируем в папку dest_dir_2 и
    вносим запись в лог-файл с действием "файл скопирован"
    1.3.2.2) если файлов .pdf обнаружено больше одного, то проверяется дата
    изменения файлов и копируется более новый, вносим запись в лог-файл с
    действием "файл скопирован"
    1.3.2.3) если год 2025, но файл в директории doc25_dir / 7. 2
    отсутствует, вносим запись в лог-файл с действием "файл не обнаружен"
    1.4) логирование ведется в едином файле, в директории dest_dir_1
    1.5) если файл 2 не обнаружен в папке doc25_dir / 7. 2, в папке
    doc24_dir / 7. 2 не проверяется наличие файла
    1.6) если в лог-файле, в строке с определенным ID, уже есть запись "Не
    найден файл .PDF", то при копировании обнаруженного нового файла, запись в
    лог-файл с действием "Файл скопирован"
    1.7) если в лог-файле, в строке с определенным ID, уже есть запись "Файл
    скопирован", но в папке обнаружен файл с новым именем (по идее, это более
    новая версия), новый файл копируется, запись в лог-файл с действием "Файл
    обновлен"

    Основная функция:
        copy_ktd_button

    На входе copy_ktd_button получаем:
        doc24_dir - root_folder_files24
        doc25_dir - root_folder_files25
        dest_dir_1 - destination_folder_1
        dest_dir_2 - destination_folder_2

    Вспомогательные функции для работы copy_ktd_button:
        find_pdf_files
        get_most_recent_file
        process_file
    """

    # Копирование .xlsb
    shutil.copy(r'\\!_Data')

    # Конвертируем .xlsb в .xlsx
    input_path = (r'\\!_Data\!2025.xlsb')
    output_path = (r'\\!_Data\!2025.xlsx')
    sheet_name = 'customer'
    date_spec = [(20, 24), (26, 35), (37, 46), (51, 58), 60, 62, 64, (66, 68),
                 (70, 71), (73, 77), (79, 82), (87, 89), (91, 93), (95, 114)]

    # Подготовка списка всех столбцов с датами
    date_columns = []
    for item in date_spec:
        if isinstance(item, tuple):
            start, end = item
            date_columns.extend(range(start, end + 1))  # + последний столбец
        else:
            date_columns.append(item)
    date_columns = sorted(set(date_columns))  # Удаляем дубликаты и сортируем

    # Чтение исходного файла
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        with open_workbook(input_path) as wb:
            with wb.get_sheet(sheet_name) as sheet:
                data = [[item.v for item in row] for row in sheet.rows()]

    headers = data[0] if len(data) > 0 else []

    # Создание нового файла с форматированием
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    # Копирование данных
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)

            # Применяем формат даты к указанным столбцам
            if col_idx in date_columns:
                cell.number_format = "dd.mm.yyyy"  # Формат даты

                # Если значение строковое, пробуем преобразовать в дату
                if isinstance(value, str):
                    try:
                        dt_value = datetime.strptime(value, "%d.%m.%Y")
                        cell.value = dt_value
                    except ValueError:
                        pass

    # Автоматическая настройка ширины столбцов, максимальная ширина - 15
    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)

        for cell in sheet[col_letter]:
            cell_len = len(str(cell.value))
            if cell_len > max_len:
                max_len = cell_len

        sheet.column_dimensions[col_letter].width = min(max_len + 2, 15)

    # Сохранение файла 2025.xlsx
    workbook.save(output_path)

    # Пути к исходной и целевой папкам, для обновления реестров 1 и 2
    source_folder = Path(r'\\Docs')
    destination_folder = Path(r'\\!_Data')

    # Список файлов для копирования
    files_to_copy = [
        'Реестр 1.xlsx',
        'Реестр 2.xlsx']

    # Копируем файлы
    for file_name in files_to_copy:
        shutil.copy(source_folder / file_name, destination_folder / file_name)

    # Конфигурация столбцов для сбора данных из файлов (номера столбцов с 0)
    config = {
        'Реестр 1': {
            'sheet_name': 'Лист1',
            'skip_rows': 2,
            'columns': {
                'number': 4,
                'date': 30
            }
        },
        'Реестр 2': {
            'sheet_name': 'Лист1',
            'skip_rows': 2,
            'columns': {
                'number': 3,
                'date1': 30,
                'date2': 50
            }
        },
        '!Трекер': {
            'sheet_name': 'customer',
            'skip_rows': 1,
            'columns': {
                'number': 0,
                'org': 46
            }
        }
    }

    # Чтение данных
    folder = Path(r'\\!_Data')

    data_p = read_sheet_data(
        folder / 'Реестр 1.xlsx',
        sheet_name=config['Реестр 1']['sheet_name'],
        skip_rows=config['Реестр 1']['skip_rows'],
        columns_config=config['Реестр 1']['columns']
    )

    data_pr = read_sheet_data(
        folder / 'Реестр 2.xlsx',
        sheet_name=config['Реестр 2']['sheet_name'],
        skip_rows=config['Реестр 2']['skip_rows'],
        columns_config=config['Реестр 2']['columns']
    )

    data_t = read_sheet_data(
        folder / '!Трекер.xlsx',
        sheet_name=config['!Трекер']['sheet_name'],
        skip_rows=config['!Трекер']['skip_rows'],
        columns_config=config['!Трекер']['columns']
    )

    # Сбор всех уникальных номеров
    all_numbers = set()
    for dataset in [data_p, data_pr, data_t]:
        for row in dataset:
            num = row.get('number')
            if num is not None:
                all_numbers.add(num)

    sorted_numbers = sorted(all_numbers, key=lambda x: str(x))

    # Создание новой таблицы
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Выборка"

    # Заголовки
    headers = [
        'номер',
        'Дата 1',
        'Дата 2',
        'организация']
    ws_out.append(headers)

    # Форматирование заголовков
    bold_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws_out.cell(row=1, column=col).font = bold_font

    # Заполнение данных
    for number in sorted_numbers:
        date_p = None
        date_pr = None
        org = None

        # Поиск в Реестре 1
        for row in data_p:
            if row['number'] == number and row['date'] is not None:
                date_p = row['date']
                break

        # Поиск в Реестре 2
        for row in data_pr:
            if row['number'] == number:
                # Берем первую непустую дату из двух столбцов
                date_pr = row.get('date1') or row.get('date2')
                if date_pr is not None:
                    break

        # Поиск в Трекере
        for row in data_t:
            if row['number'] == number and row.get('org') is not None:
                org = row['org']
                break

        ws_out.append([number, date_p, date_pr, org])

    # Настройка ширины столбцов
    for col_letter, width in {'A': 7, 'B': 21, 'C': 28, 'D': 13}.items():
        ws_out.column_dimensions[col_letter].width = width

    # Форматирование дат
    date_format = 'dd.mm.yyyy'  # Короткий формат даты
    for row in ws_out.iter_rows(min_row=2, max_col=3, max_row=ws_out.max_row):
        for cell in row[1:3]:  # Столбцы B и C
            if isinstance(cell.value, datetime):  # Проверяем, что это дата
                cell.number_format = date_format

    # Сохранение файла, при необходимости
    # wb_out.save(folder / '!_Выборка из трекера и реестров.xlsx')

    # Проверка наличия лог-файла и его создание, если не существует
    log_path = os.path.join(dest_dir_1, 'log.xlsx')
    if not os.path.exists(log_path):
        create_log_file(log_path)

    # Открытие лог-файла для добавления записей
    log_wb = load_workbook(log_path)

    # Чтение таблицы Excel, начиная со второй строки
    rows = list(wb_out['Выборка'].iter_rows(min_row=2, values_only=True))
    wb_out.close()

    for row in rows:
        current_id = row[0]  # ID
        date_ready = row[1] if len(row) >= 2 else None  # Дата согл. 1
        date_approval = row[2] if len(row) >= 3 else None  # Дата согл. 2
        organization = row[3] if len(row) >= 4 else ''  # Организация

        # Проверка наличия обязательных полей
        if current_id is None or (not date_ready and not date_approval):
            continue

        # Подбираем нужную дату
        selected_date = date_ready if date_ready else date_approval
        # Пропускаем строку, если дата некорректная
        if not isinstance(selected_date, datetime):
            continue

        # Проверяем допустимые годы обработки
        year = selected_date.year
        if year not in [2024, 2025]:
            continue

        # Определение 1 или 2 и выбор директории для копирования
        is_rd = bool(date_ready)
        subfolder = '1' if is_rd else '2'
        destination_folder = dest_dir_1 if is_rd else dest_dir_2

        # Работа с файлом
        match year:
            case 2024:
                search_folder = os.path.join(
                    doc24_dir, f"{int(current_id):04}", subfolder)
                found_pdfs = find_pdf_files(search_folder)
                if found_pdfs:
                    file_to_copy = get_most_recent_file(
                        found_pdfs, search_folder)
                    process_file(
                        file_to_copy, search_folder, current_id, subfolder,
                        destination_folder, log_wb, organization, dest_dir_1)
                else:
                    log_wb.active.append([
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'Не найден файл 2024.PDF в папке 2024',
                        current_id,
                        '',
                        subfolder.split(' ')[1]])

            case 2025:
                # 1. Проверяем в doc25_dir с точным совпадением номера (2025)
                search_folder = os.path.join(
                    doc25_dir, f"{int(current_id):04}", subfolder)
                found_pdfs = find_pdf_files(search_folder)

                if not found_pdfs:
                    # 2. Если нет, проверяем с доп. символом (2025)
                    search_pattern = os.path.join(
                        doc25_dir, f"{int(current_id):04}*", subfolder)
                    possible_folders = [
                        f for f in glob(search_pattern) if f != os.path.join(
                            doc25_dir, f"{int(current_id):04}", subfolder)]
                    if possible_folders:
                        search_folder = possible_folders[0]
                        found_pdfs = find_pdf_files(search_folder)

                if not found_pdfs:
                    # 3. Если нет, проверяем с точным совпадением номера (2024)
                    search_folder = os.path.join(
                        doc24_dir, f"{int(current_id):04}", subfolder)
                    found_pdfs = find_pdf_files(search_folder)

                if not found_pdfs:
                    # 4. Если нет, проверяем с доп. символом (2024)
                    search_pattern = os.path.join(
                        doc24_dir, f"{int(current_id):04}*", subfolder)
                    possible_folders = [
                        f for f in glob(search_pattern) if f != os.path.join(
                            doc24_dir, f"{int(current_id):04}", subfolder)]
                    if possible_folders:
                        search_folder = possible_folders[0]
                        found_pdfs = find_pdf_files(search_folder)

                # Обработка результата поиска
                if found_pdfs:
                    file_to_copy = get_most_recent_file(
                        found_pdfs, search_folder)
                    # Проверяем соответствие имени файла шаблону
                    if "2024" in search_folder and not re.fullmatch(
                            r'\d{6}(2025|25).*\.pdf', file_to_copy):
                        log_wb.active.append([
                            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'Не найден файл 2025.PDF в папке 2024',
                            current_id, '', subfolder.split(' ')[1]])
                    else:
                        process_file(
                            file_to_copy, search_folder, current_id,
                            subfolder, destination_folder, log_wb,
                            organization, dest_dir_1)
                else:
                    log_wb.active.append([
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'Не найден файл 2025.PDF в папке 2025',
                        current_id, '', subfolder.split(' ')[1]])

    # Добавляем строку в качестве разделителя и заканчиваем работу с лог-файлом
    log_wb.active.append(['работа завершена'])
    log_wb.save(log_path)
    log_wb.close()

    # Открываем директорию назначения (возможно открытие файла)
    open_file_and_folder(os.path.join(dest_dir_1, 'log.xlsx'))


# Функция для кнопки 5 - указание пути к лог-файлу по 1 и 2
def open_log_file_rd(content_df_1):
    """
    На входе, функция получает:
    - content_df_1 (destination_folder_1) - директория назначения для 1
    """
    # Формируем путь к лог-файлу
    log_file_path = os.path.join(content_df_1, 'log.xlsx')
    # Передаем путь к директории в функцию
    check_and_open_log(log_file_path)


# Вспомогательная функция для copy_ktd_button - считывание данных из файлов
def read_sheet_data(path, sheet_name, skip_rows=0, columns_config=None):
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet_name]

    row_iter = ws.iter_rows()

    # Пропускаем заданное количество строк
    for _ in range(skip_rows):
        next(row_iter)

    data = []
    for row_cells in row_iter:
        row_data = {}
        for key, col_idx in columns_config.items():
            if col_idx < len(row_cells):
                value = row_cells[col_idx].value
                # Преобразуем строки в даты, если это столбец с датой
                if key in (
                        'date', 'date1', 'date2') and isinstance(value, str):
                    try:
                        value = datetime.strptime(value, '%d.%m.%Y')
                    except ValueError:
                        value = None  # Если преобразование не удалось - None
                row_data[key] = value
            else:
                row_data[key] = None
        data.append(row_data)

    wb.close()
    return data


# Вспомогательная функция для copy_ktd_button - поиск PDF-файлов в директории
def find_pdf_files(doc_dir):
    return [f for f in os.listdir(doc_dir) if f.lower().endswith(
        '.pdf')] if os.path.exists(doc_dir) else []


# Вспом. функция для copy_ktd_button - выбор самого нового файла из списка
def get_most_recent_file(file_list, search_folder):
    return max(file_list, key=lambda f: os.path.getmtime(
        os.path.join(search_folder, f)), default=None)


# Вспом. функция для copy_ktd_button - работа с PDF-файлом и логирование
def process_file(file_to_copy, search_folder, current_id, subfolder,
                 destination_folder, log_wb, org_dir, dest_dir_log):

    not_org_dir = os.path.join(destination_folder, 'Не распределены')
    not_org_file = os.path.join(
        not_org_dir, f"{int(current_id):04} - {file_to_copy}")

    # Если файл уже есть в "Не распределенных", но появилась новая информация
    if org_dir in (('contractor_1', 'contractor_2') and
                   os.path.exists(not_org_file)):
        try:
            target_dir = os.path.join(destination_folder, org_dir)
            os.makedirs(target_dir, exist_ok=True)

            # Перемещаем файл в соответствующую папку
            shutil.move(not_org_file, os.path.join(
                target_dir, os.path.basename(not_org_file)))

            log_wb.active.append([
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                f"Перемещен из 'Не распределены' в '{org_dir}'",
                current_id,
                file_to_copy,
                subfolder.split(' ')[1],
                org_dir
            ])

            return True
        except Exception as e:
            log_wb.active.append([
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                f'Ошибка перемещения: {str(e)}',
                current_id,
                file_to_copy
            ])
            return False

    target_dir = os.path.join(
        destination_folder, 'Не распределены' if org_dir not in (
            'contractor_1', 'contractor_2') else org_dir)
    os.makedirs(target_dir, exist_ok=True)
    src_file = os.path.join(search_folder, file_to_copy)
    new_filename = f"{int(current_id):04} - {file_to_copy}"
    dest_file = os.path.join(target_dir, new_filename)

    try:
        # Проверка лог-файла
        log_path = os.path.join(dest_dir_log, 'log.xlsx')
        logged_files = set()  # Множество для хранения уже скопированных файлов
        logged_ids = {}  # Словарь для хранения ID и их статусов
        if os.path.exists(log_path):
            wb = load_workbook(log_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                logged_files.add(row[3])  # Имя файла из четвертого столбца
                # Номер: (результат, имя файла)
                logged_ids[row[2]] = (row[1], row[3])
            wb.close()

        # Если файл уже скопирован, пропускаем
        if file_to_copy in logged_files:
            return True

        # Определяем сообщение для лога
        if current_id not in logged_ids:
            log_message = 'Файл скопирован'
        elif logged_ids[current_id][0] == 'Не найден файл .PDF':
            log_message = 'Файл скопирован'
        elif logged_ids[current_id][1]:  # Если ранее был скопирован файл
            log_message = 'Файл обновлен'
        else:
            log_message = 'Файл скопирован'

        # Копируем файл и вносим результат в лог-файл
        shutil.copy2(src_file, dest_file)
        log_wb.active.append([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            log_message,
            current_id,
            file_to_copy,
            subfolder.split(' ')[1],
            os.path.basename(os.path.dirname(dest_file))
        ])
        return True
    except Exception as e:
        log_wb.active.append([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            f'Ошибка копирования: {str(e)}',
            current_id,
            file_to_copy
        ])
        return False


# Функция создания лог-файла (универсальная)
def create_log_file(log_path):
    wb = Workbook()
    ws = wb.active
    ws.append([
        'Дата', 'Результат', 'Номер', 'Имя файла',
        'Тип документа', 'Куда скопирован'])

    # Настройка стиля заголовка
    for cell in ws[1]:  # Первая строка (заголовок)
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='thick'))

    # Установка ширины столбцов
    for col_letter, width in {
            'A': 19, 'B': 47, 'C': 6, 'D': 100, 'E': 15, 'F': 22}.items():
        ws.column_dimensions[col_letter].width = width

    # Сохранение файла
    wb.save(log_path)


# Функция для открытия лог-файла (универсальная)
def check_and_open_log(log_file_path):
    """
    На входе, функция получает:
    - log_file_path - путь к директории с лог-файлом
    """
    # Проверяем, существует ли файл
    if os.path.exists(log_file_path):
        try:
            # Открываем файл в зависимости от операционной системы
            if os.name == 'nt':  # Для Windows
                os.startfile(log_file_path)
            # elif os.name == 'posix':  # Для macOS и Linux
            #     subprocess.run(['open', log_file_path] if
            #       sys.platform == 'darwin' else ['xdg-open', log_file_path])
        except Exception as e:
            messagebox.showerror('Ошибка', f'Не удалось открыть файл: {e}')
    else:
        messagebox.showwarning(
            'Предупреждение', f'Лог-файл не найден по пути: {log_file_path}')


# Функция для открытия файла и директории (универсальная)
def open_file_and_folder(output_filename):
    """
    На входе, функция получает:
    - output_filename - имя файла, включая путь до него

    Возможно так же, открытие сформированного файла, для этого разблокировать
    соответствующие строки ниже.
    """
    if os.name == 'nt':  # Для Windows - файл, директория
        # os.startfile(output_filename)  # Открываем файл
        os.startfile(os.path.dirname(output_filename))  # Открываем директорию
    # elif os.name == 'posix':  # Для macOS и Linux
    #     # Для macOS - файл, директория
    #     subprocess.run(['open', output_filename])
    #     subprocess.run(['open', os.path.dirname(output_filename)])

    # # Для Linux - файл, директория
    # subprocess.run(['xdg-open', output_filename])
    # subprocess.run(['xdg-open', os.path.dirname(output_filename)])


# Функция для конвертации file_2 в список
def transform_excel_list(input_fn, output_file):
    """
    На входе, функция получает:
    - input_fn (filename) - имя исходного файла, включая путь до него
    - output_file - имя выходного файла, включая путь до него
    """
    wb_out = None  # Инициализация переменной перед try-блоком

    # Проверяем наличие выходного файла и создаем его, если его нет
    try:
        wb_out = load_workbook(output_file)
    except FileNotFoundError:
        wb_out = Workbook()
        wb_out.save(output_file)
    finally:
        if wb_out is not None:
            wb_out.close()

    # Открываем исходную книгу
    wb_in = load_workbook(input_fn)
    ws_in = wb_in.active

    # Открываем новую книгу для записи
    wb_out = load_workbook(output_file)
    ws_out = wb_out.active

    # Определяем триггеры для обрезки
    trigger_start = '№ п/п'
    trigger_end = 'Исполнитель'

    # Инициализируем счетчики для индексов
    row_start = 0
    row_end = ws_in.max_row

    # Проходим по таблице и находим строки с триггерами
    for row_num in range(1, ws_in.max_row + 1):
        for col_num in range(1, ws_in.max_column + 1):
            cell_value = ws_in.cell(row=row_num, column=col_num).value

            if cell_value == trigger_start:
                # Начинаем копирование со строки с триггером
                row_start = row_num
            elif cell_value == trigger_end:
                # Заканчиваем копирование на две строки выше строки с триггером
                row_end = row_num - 2
                break

    # Копируем строки между найденными границами
    for row_num in range(row_start, row_end + 1):
        for col_num in range(1, ws_in.max_column + 1):
            source_cell = ws_in.cell(row=row_num, column=col_num)
            dest_cell = ws_out.cell(
                row=row_num - row_start + 1, column=col_num)
            dest_cell.value = source_cell.value

            # Копируем стиль ячейки
            if source_cell.has_style:
                dest_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic
                )  # Шрифт
                dest_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )  # Границы ячейки
                dest_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )  # Стиль содержимого ячейки

    # Сохраняем промежуточный файл
    wb_out.save(output_file)

    # Теперь открываем файл снова и удаляем пустые столбцы и строки
    wb_out = load_workbook(output_file)
    ws_out = wb_out.active

    # Перебираем все столбцы и проверяем, заполнены ли они
    empty_columns = []
    for col_idx in range(1, ws_out.max_column + 1):
        empty = True
        for row_idx in range(1, ws_out.max_row + 1):
            if ws_out.cell(row=row_idx, column=col_idx).value is not None:
                empty = False
                break
        if empty:
            empty_columns.append(col_idx)

    # Перебираем все строки и проверяем, заполнены ли они
    empty_rows = []
    for row_idx in range(1, ws_out.max_row + 1):
        empty = True
        for col_idx in range(1, ws_out.max_column + 1):
            if ws_out.cell(row=row_idx, column=col_idx).value is not None:
                empty = False
                break
        if empty:
            empty_rows.append(row_idx)

    # Удаляем пустые столбцы
    for col_idx in sorted(empty_columns, reverse=True):
        ws_out.delete_cols(col_idx, 1)

    # Удаляем пустые строки
    for row_idx in sorted(empty_rows, reverse=True):
        ws_out.delete_rows(row_idx, 1)

    # Задаем ширину определенных столбцов
    column_widths = {
        'A': 6,
        'B': 65,
        'C': 150,
        'D': 12,
        'E': 12
    }

    for col_name, width in column_widths.items():
        ws_out.column_dimensions[col_name].width = width

    # Задаем высоту строк в зависимости от наличия текста
    for row in ws_out.iter_rows():
        max_lines = 1  # Минимальное значение

        for cell in row:
            if cell.value and isinstance(cell.value, str):

                # Расчет учитывает длину текста и ширину столбца
                col_width = ws_out.column_dimensions[
                    get_column_letter(cell.column)].width
                approx_chars_per_line = col_width / 2  # Эмп. коэф.
                text_length = len(cell.value)
                length_based_lines = (
                    (text_length / 1.3) / approx_chars_per_line)
                # 2 - эмпирический коэффициент (примерное кол-во символов на
                # единицу ширины), можно регулировать

                # Берем максимальное значение
                current_lines = max(max_lines, length_based_lines)
                if current_lines > max_lines:
                    max_lines = current_lines

        # Устанавливаем высоту (13 единиц на строку, можно регулировать)
        if max_lines > 1:
            ws_out.row_dimensions[row[0].row].height = 13 * max_lines

    # Сохраняем промежуточный файл
    wb_out.save(output_file)

    # Чистим третий столбец от лишних пробелов, табуляции и переносов строк
    workbook = load_workbook(output_file)
    sheet = workbook.active  # получаем активный лист (первый в файле)

    # Проходим по всем ячейкам в указанном столбце
    for row in sheet.iter_rows(min_col=3, max_col=3):
        cell = row[0]  # В каждой строке берем только ячейку 3-го столбца
        if isinstance(cell.value, str):
            cell.value = re.sub(r'\s+', ' ', cell.value.strip())

    # Сохраняем изменения
    workbook.save(output_file)
    wb_in.close()
    wb_out.close()


# Формирование файла file_2, на основе файла file_1
def transform_excel_removal_equipment(input_fn, output_file, content_fwt_rrp):
    """
    На входе, функция получает:
    - input_fn - имя исходного файла, включая путь до него
    - output_file - имя выходного файла, включая путь до него
    - content_fwt_rrp - путь к информации
    """

    # Функция для изменения содержимого объединенной ячейки
    def set_merged_cell(cell_ref_content, value_content):
        cell_content = ws[cell_ref_content]
        cell_content.value = value_content

    # Настройка стилей первой группы ячеек
    def merge_and_style_1(ws_mas_1, merged_range_mas_1):
        # Настраиваем стили
        font = Font(name='Times New Roman', size=10, bold=True)
        alignment_mas_1 = Alignment(horizontal='center', vertical='center')

        # Объединяем ячейки
        ws_mas_1.merge_cells(merged_range_mas_1)

        # Получаем главную ячейку (первую в диапазоне)
        start_cell = merged_range_mas_1.split(':')[0]
        main_cell_1 = ws_mas_1[start_cell]

        # Применяем стили
        main_cell_1.font = font
        main_cell_1.alignment = alignment_mas_1

    # Настройка стилей второй группы ячеек
    def merge_and_style_2(ws_mas_2, merged_range_mas_2):
        # Настраиваем стили
        font = Font(name='Times New Roman', size=7)
        alignment_mas_2 = Alignment(horizontal='center', vertical='center')
        top_border = Border(top=Side(style='thin'))

        # Объединяем ячейки
        ws_mas_2.merge_cells(merged_range_mas_2)

        # Получаем главную ячейку (первую в диапазоне)
        start_cell = merged_range_mas_2.split(':')[0]
        main_cell_2 = ws_mas_2[start_cell]

        # Применяем стили
        main_cell_2.font = font
        main_cell_2.alignment = alignment_mas_2

        # Получаем границы диапазона
        min_col_mas_2, min_row_mas_2, max_col_mas_2, _ = (
            range_boundaries(merged_range_mas_2))

        # Применяем отображение верхней границы
        for col in range(min_col_mas_2, max_col_mas_2 + 1):
            cell_mas_2 = ws.cell(row=min_row_mas_2, column=col)
            cell_mas_2.border = top_border

    # Обновляет объединенные ячейки
    def process_merged_ranges(
            ws_pmr, merged_ranges_pmr, value_pmr, alignment_pmr):
        for merged_range_pmr in merged_ranges_pmr:
            min_col, min_row, _, _ = range_boundaries(merged_range_pmr)
            main_cell = ws_pmr.cell(row=min_row, column=min_col)
            main_cell.value = value_pmr
            main_cell.alignment = alignment_pmr

    # 1) Загружаем файл
    wb = load_workbook(input_fn)
    sheet_name = 'Лист_1'  # Название листа
    ws = wb[sheet_name]

    # 2) Корректируем строку с первой позицией ведомости.
    # Настройки выравнивания и отображения текста
    alignment = Alignment(
        wrap_text=True,  # Отображения длинного текста в нескольких строках
        horizontal='center',  # Включение выравнивания
        vertical='center'  # Включение выравнивания
    )

    # Список объединенных диапазонов
    merged_ranges = [
        'D27:R27',
        'S27:Z27',
        'AA27:AE27'
    ]

    # Значение для всех объединенных ячеек
    target_value_null = '0'
    target_value_edit = ''

    # Название листа в Excel файле
    sheet_name = 'Выгрузка'

    # Индексы столбцов (начиная с 1)
    number_column_index = 2  # Первый столбец с числами
    data_column_index = 114  # Второй столбец с данными

    # Название, которое мы ищем в столбце с данными
    target_content = 'не требуется'

    # Загружаем Excel файл
    workbook = load_workbook(content_fwt_rrp)
    sheet = workbook[sheet_name]

    # Проходим по строкам в листе
    file_name = os.path.splitext(os.path.basename(input_fn))[0].zfill(4)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        number = row[number_column_index - 1]  # Данные из первого столбца
        data = row[data_column_index - 1]  # Данные из второго столбца

        # Проверяем, содержит ли строка искомое название
        if data is not None and target_content in str(data):
            # Преобразуем число в строку и дополняем нулями до 4 символов
            number_str = str(number).zfill(4) if number is not None else "0000"

            # Проверяем, начинаются ли первые 4 символа имени файла с числа
            if file_name == number_str:
                process_merged_ranges(
                    ws, merged_ranges, target_value_null, alignment)
                set_merged_cell(
                    'AF27', 'не предусмотрен')

                output_file = Path(output_file)
                output_file = (
                    output_file.parent/"Для customer"/output_file.name)
        else:
            # Преобразуем число в строку и дополняем нулями до 4 символов
            number_str = str(number).zfill(4) if number is not None else "0000"

            # Проверяем, начинаются ли первые 4 символа имени файла с числа
            if file_name == number_str:
                process_merged_ranges(
                    ws, merged_ranges, target_value_edit, alignment)
                set_merged_cell('AF27', '')

    # Включение отображения длинного текста в ячейке AF27
    cell = ws['AF27']
    cell.alignment = alignment

    # 3) Определяем диапазон очистки: со строки 28 и до конца листа
    clear_min_row = 28
    clear_max_row = ws.max_row

    # Удаляем (разъединяем) объединенные ячейки в диапазоне
    merged_ranges = ws.merged_cells.ranges.copy()
    for merged_range in merged_ranges:
        first_row = merged_range.min_row
        last_row = merged_range.max_row
        if first_row >= clear_min_row and last_row <= clear_max_row:
            ws.unmerge_cells(str(merged_range))

    # Очищаем данные и форматирование в диапазоне
    default_font = Font(
        name='Times New Roman', size=11, bold=False, italic=False)
    default_fill = PatternFill(fill_type=None)
    default_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    default_alignment = Alignment(horizontal='general', vertical='bottom')

    for row in ws.iter_rows(min_row=clear_min_row, max_row=clear_max_row):
        for cell in row:
            cell.value = None
            cell.font = default_font
            cell.fill = default_fill
            cell.border = default_border
            cell.alignment = default_alignment

    # Сбрасываем высоту строк, в которых была очистка данных
    for row in range(clear_min_row, clear_max_row + 1):
        if row in ws.row_dimensions:
            ws.row_dimensions[row].height = None

    # 4) Настраиваем и заполняем блок подписантов

    # Список диапазонов для объединения и стилизации первой группы ячеек
    merged_ranges_group_1 = [
        'B30:F30',
        'B35:F35',
        'G30:K30',
        'G35:K35',
        'X30:AD30',
        'X35:AD35',
        'AF30:AP30',
        'AF35:AP35'
    ]

    # Применяем функцию к каждому диапазону
    for merged_range_1 in merged_ranges_group_1:
        merge_and_style_1(ws, merged_range_1)

    # Список диапазонов для объединения и стилизации второй группы ячеек
    merged_ranges_group_2 = [
        'G31:K31',
        'G36:K36',
        'X31:AD31',
        'X36:AD36',
        'AF31:AP31',
        'AF36:AP36'
    ]

    # Применяем функцию к каждому диапазону
    for merged_range_2 in merged_ranges_group_2:
        merge_and_style_2(ws, merged_range_2)

        # Список ячеек и значений
        cells_values = [
            ('G22', 'ВЕДОМОСТЬ'),
            ('B30', 'Исполнитель'),
            ('G31', '(должность)'),
            ('X31', '(подпись)'),
            ('AF31', '(расшифровка подписи)'),
            ('B35', 'Заказчик'),
            ('G36', '(должность)'),
            ('X36', '(подпись)'),
            ('AF36', '(расшифровка подписи)')
        ]

        # Вносим данные
        for cell_ref, value in cells_values:
            set_merged_cell(cell_ref, value)

    # 5) Устанавливаем высоты строк первой позиции и блока подписантов
    row_heights = {
        27: 50,
        30: 33,
        35: 33
    }

    for row_num, height in row_heights.items():
        ws.row_dimensions[row_num].height = height

    # Сохраняем изменения в файл
    wb.save(output_file)


# Пути к информации
folder_with_tr_r = None
folder_with_tr_2025 = None

# Директория с file_1 по умолчанию
default_folder_file_1 = None

# Исходные директории с файлами
root_folder_k24 = None
root_folder_k25 = None

# Целевые директории для 1 и 2
destination_folder_1 = None
destination_folder_2 = None

# Создание главного окна
root = tk.Tk()
root.title('The F.I.R. (File Interaction Runner)')

# Установка размеров окна
root.geometry('350x215')

# Текст над полем ввода
label1 = tk.Label(
    root,
    text='Работаем с папкой (изменить при необходимости):',
    font=('Arial', 10))
label1.place(x=10, y=5)

# Поле для ввода текста (Entry)
entry = tk.Entry(root, width=54)
entry.place(x=10, y=30)

# Установка значения по умолчанию
entry.insert(0, default_folder_file_1)  # Вставляем значение по умолчанию

# Кнопка 1 - Перемещение файлов в целевые директории и создание списка
button1 = tk.Button(
    root,
    text='Обработать файлы в папке',
    command=lambda: copy_file_1_button(
        folder_with_tr_r,
        folder_with_tr_2025)
)
button1.place(x=10, y=55, width=235, height=30)

# Кнопка 2 - Открытие лог-файла
button2 = tk.Button(
    root,
    text='Открыть лог-файл',
    command=open_log_file_file_1
)
button2.place(x=10, y=90, width=235, height=30)

# Кнопка 3 - Создание из выбранного файла, списка и файла сверки
button3 = tk.Button(
    root,
    text='Сформировать список',
    command=open_file_file_2
)
button3.place(x=10, y=125, width=235, height=30)

# Надпись над текстовым полем для отображения обработанных файлов
processed_files_label = tk.Label(root, text='Обработаны:', font=('Arial', 10))
processed_files_label.place(x=255, y=50)

# Текстовое поле для отображения обработанных файлов
text_area = tk.Text(root, width=10, height=5)
text_area.place(x=255, y=70)

# Кнопка 4 - Поиск и копирование 1
button4 = tk.Button(
    root,
    text='Скопировать новые файлы',
    command=lambda: copy_ktd_button(
        root_folder_k24,
        root_folder_k25,
        destination_folder_1,
        destination_folder_2
    )
)
button4.place(x=10, y=175, width=250, height=30)

# Кнопка 5 - Открытие лога
button5 = tk.Button(
    root,
    text='Открыть лог',
    command=lambda: open_log_file_rd(destination_folder_1)
)
button5.place(x=265, y=175, width=75, height=30)

# Запуск главного цикла обработки событий
root.mainloop()
