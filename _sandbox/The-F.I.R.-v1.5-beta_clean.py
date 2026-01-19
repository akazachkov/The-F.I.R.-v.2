import re
import shutil
import tkinter as tk
import openpyxl
import warnings  # Подавление предупреждающих сообщений, при работе с .xlsb
import platform
import subprocess  # nosec B404
from pyxlsb import open_workbook
from datetime import datetime
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl import load_workbook, Workbook
from tkinter import messagebox, filedialog
from pathlib import Path
from glob import glob
import sys  # Принудительное отображение в терминале, кодировки utf-8


# Подавление ошибки, связанной с ограниченной поддержкой библиотеки openpyxl,
# функций проверки данных в Excel
warnings.filterwarnings(
    "ignore", category=UserWarning,
    message="Data Validation extension is not supported")

# Принудительное отображение в терминале, кодировки utf-8
sys.stdout.reconfigure(encoding='utf-8')

"""
The F.I.R. (File Interaction Runner) v1.5 (30.06.2025)

beta-версия, которая так и пошла в общее пользование из-за отсутствия времени
на полноценное тестирование и доработку. Как и в случае с v1.4.1 - при
"стерилизации", приложение перестало работать.
Сохранено только ради возможности вспомнить "как что было реализовано" и
возможно, забрать какие-то блоки кода или задумки, в "The F.I.R. v.2".

Основные изменения версии:
- Обновление в работе с файлами - если contractor_2, то в целевую папку "Для
customer", копируется файл.
- Копирование файлов вынесено из блока работы в отдельный функционал.
- Переработаны функции открытия лог-файлов и папок с файлами.
- Исключено использование библиотеки os. Все действия с путями и именами
файлов, теперь выполняются с помощью библиотеки pathlib.
- Код приведён в соответствие стандарту PEP8 и устранены наиболее важные
предупреждения Bandit. B404, B603 и B607 в определённых строках намеренно
заглушены.
- Добавлены новые кнопки:
1) копирование в папку трекера, реестров и т.д.;
2) открытие папки.
"""


# Функция для кнопки копирования рабочих файлов из разных источников
def copy_file_for_data(
        content_folder_data,
        content_ip_folder,
        content_source_folder_dp,
        content_root_folder_kd,
        content_ofnt,
        content_fnr_t,
        content_fnr_2,
        content_fnr_1,
        content_customer):
    '''
    content_folder_data - целевая папка
    content_customer - наименование заказчика

    Папки с исходными файлами:
    content_ip_folder
    content_source_folder_dp
    content_root_folder_kd

    Исходные (оригинальные) имена файлов:
    content_ofnt
    content_fnr_t
    content_fnr_1
    content_fnr_2
    '''
    # Источники - папка_1: [файл_1, файл_2], папка_2: [файл_3, файл_4] и т.д.
    source_folders = {
        content_source_folder_dp: [
            content_ofnt,
            content_fnr_t],
        content_root_folder_kd: [
            content_fnr_2,
            content_fnr_1]}

    # Копируем в целевую папку файлы с определёнными именами
    for source, files in source_folders.items():
        for file in files:
            shutil.copy(source / file, content_folder_data / file)

    # Удаляем из целевой папки старый файл с "план" в имени
    for old_file in content_folder_data.glob('*'):
        if old_file.is_file() and "план" in old_file.name.lower():
            old_file.unlink()

    # Ищем в исходной папке файл с "план" в имени и копируем в целевую папку
    file_found_I = None
    for file in content_ip_folder.iterdir():
        if file.is_file() and "план" in file.name.lower():
            file_found_I = file
            break

    shutil.copy(file_found_I, content_folder_data / file_found_I.name)

    # Конвертируем файл .xlsb в .xlsx
    input_path = (content_folder_data / content_ofnt)
    output_path = (
        (content_folder_data / content_ofnt).with_suffix(".xlsx"))

    # Столбцы с информацией в формате дат
    date_spec = [(20, 24), (26, 35), (37, 46), (51, 58), 60, 62, 64, (66, 68),
                 (70, 71), (73, 77), (79, 82), (87, 89), (91, 93), (95, 114)]

    # Подготовка списка всех столбцов с датами
    date_columns = []
    for item in date_spec:
        if isinstance(item, tuple):
            start, end = item
            date_columns.extend(range(start, end + 1))  # + последний столбец
        else:
            date_columns.append(item)
    date_columns = sorted(set(date_columns))  # Удаляем дубликаты и сортируем

    # Чтение исходного файла
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        with open_workbook(input_path) as wb:
            with wb.get_sheet(content_customer) as sheet:
                data = [[item.v for item in row] for row in sheet.rows()]

    headers = data[0] if len(data) > 0 else []

    # Создание нового файла с форматированием
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = content_customer

    # Копирование данных
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)

            # Применяем формат даты к указанным столбцам
            if col_idx in date_columns:
                cell.number_format = "dd.mm.yyyy"  # Формат даты

                # Если значение строковое, пробуем преобразовать в дату
                if isinstance(value, str):
                    try:
                        dt_value = datetime.strptime(value, "%d.%m.%Y")
                        cell.value = dt_value
                    except ValueError:
                        pass

    # Автоматическая настройка ширины столбцов, максимальная ширина - 15
    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)

        for cell in sheet[col_letter]:
            cell_len = len(str(cell.value))
            if cell_len > max_len:
                max_len = cell_len

        sheet.column_dimensions[col_letter].width = min(max_len + 2, 15)

    # Сохранение файла с расширением .xlsx
    workbook.save(output_path)

    messagebox.showinfo('Результат', 'Дело сделано')


# Функция для кнопки 1 - 'Обработать файлы в директории'
def copy_1_button(
        content_fwt_r,
        content_fwt,
        content_wc_2,
        content_customer,
        content_contractor_1,
        content_contractor_2,
        content_suffix_1,
        content_suffix_2,
        content_ffc):
    '''
    Пути к рабочим файлам:
    content_fwt_r
    content_fwt
    content_wc_2

    Наименования организаций:
    content_customer
    content_contractor_1
    content_contractor_2

    Приставки для итоговых файлов:
    content_suffix_1,
    content_suffix_2

    content_ffc - название подпапки
    '''
    # Получаем путь из поля ввода
    source_directory = Path(entry.get())

    # Получаем список файлов в исходной директории (без поддиректорий)
    files = [f for f in source_directory.iterdir() if f.is_file()]

    # Если файлов нет - выводим сообщение и завершаем функцию
    if not files:
        print('Файлов не обнаружено')
        return

    # Проверка наличия лог-файла и его создание, если не существует
    log_path = source_directory / '_log_1.xlsx'
    if not log_path.exists():
        create_log_file(log_path)

    # Открытие лог-файла для добавления записей
    log_wb = load_workbook(log_path)

    # Получаем базовую директорию
    base_directory = source_directory.parent

    # Загружаем таблицу для определения организации
    wb = load_workbook(content_fwt)
    ws = wb[content_customer]
    org_mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        current_id = row[0]  # Столбец - номер
        org = row[46] if len(row) > 47 else ''  # Столбец - Орг
        org_mapping[str(current_id)] = org if org in [
            content_contractor_1,
            content_contractor_2] else 'Не определено'
    wb.close()

    for file in files:
        # Определяем имя исходного файла, включая путь до него
        input_filename = source_directory / file

        # Получаем имя файла и расширение
        file_name, extension = file.stem, file.suffix

        # Пропускаем лог-файл и работаем с остальными файлами в папке
        if file == log_path:
            continue

        # Извлекаем часть до первого тире (если есть)
        if '-' in file_name:
            base_name = file_name.split('-')[0].strip()
        else:
            base_name = file_name

        # Дополняем base_name нулями до 4 символов
        file_name = base_name.zfill(4)

        # Определяем организацию
        org = org_mapping.get(file_name.lstrip('0'), 'Не определено')

        # Создаем поддиректорию "Для загрузки" в исходной директории
        upload_dir = source_directory / 'Для загрузки' / org
        upload_dir.mkdir(parents=True, exist_ok=True)

        # Копируем файл в поддиректорию "Для загрузки"
        upload_path = upload_dir / file.name

        shutil.copy(input_filename, upload_path)

        # Добавляем суффикс перед расширением
        new_file_name = f'{file_name}{content_suffix_1}{extension}'

        # Определяем путь к директории назначения
        destination_folder = base_directory / file_name

        # Проверяем существование директории назначения
        if not destination_folder.exists():
            destination_folder.mkdir(parents=True, exist_ok=True)

        # Если contractor_2, то копируем файл
        if content_wc_2.exists() and org == content_contractor_2:
            shutil.copy(content_wc_2, destination_folder / content_ffc)

        # Формируем полный путь к файлу 1 в новой директории
        destination = destination_folder / new_file_name

        # Формируем имя файла для списка
        output_file_inserted_equipment = destination_folder / f'{
            file_name} Список{extension}'
        transform_excel_list(
            input_filename,
            output_file_inserted_equipment)

        # Формируем имя файла
        output_file_removal_equipment = destination_folder / f'{
            file_name}{content_suffix_2}{extension}'
        transform_excel_removal_equipment(
            input_filename,
            output_file_removal_equipment,
            content_fwt_r)

        # Перемещаем файл в целевую директорию
        shutil.move(input_filename, destination)

        # Добавляем запись в лог-файл о завершении перемещения
        log_wb.active.append([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Обработан и перемещён\n',
            file_name,
            str(destination),
            content_suffix_1[1:],
            org])

        # Добавляем имя обработанного файла в текстовое поле
        text_area.insert(tk.END, f'{file_name}\n')

    # Добавляем строку в качестве разделителя и закрываем лог-файл
    log_wb.active.append(['работа завершена'])
    log_wb.save(log_path)
    log_wb.close()

    messagebox.showinfo('Результат', 'Дело сделано')


# Функция для кнопки 2 - указание пути к лог-файлу
def open_log_file_():
    # Получаем путь из поля ввода
    directory_path = Path(entry.get())
    # Формируем путь к лог-файлу
    log_file_path = (directory_path / '_log_1.xlsx')
    # Передаём путь к директории в функцию
    check_and_open_log(log_file_path)


# Функция для кнопки 3 - открытие окна выбора файла и формирование списков
def open_file_2():
    # Открываем диалоговое окно выбора файла
    input_filename = Path(
        filedialog.askopenfilename(
            title='Выберите файл',  # Заголовок окна
            filetypes=(('Все файлы', '*.*'), ('Файлы Excel', '*.xlsx'))))

    # Проверяем, был ли выбран файл
    if input_filename:

        # Получаем путь к директории и разбиваем имя файла на имя и расширение
        input_dir = input_filename.parent
        filename = input_filename.stem
        file_extension = input_filename.suffix

        # Создаем имя выходного файла, включая путь до него
        output_filename_1 = input_dir / f'{
            filename[:4]} Список{file_extension}'

        # Запускаем функцию, для создания списка
        transform_excel_list(input_filename, output_filename_1)

        # Открываем директорию назначения
        open_file_and_folder(output_filename_1)
    else:
        print('Файл не выбран')


# Функция для кнопки 4 - поиск и копирование 1, при отсутствии 1 - 2
def copy_ktd_button(
        doc24_dir,
        doc25_dir,
        dest_dir_1,
        dest_dir_2,
        content_folder_data,
        content_customer,
        content_contractor_1,
        content_contractor_2):
    # Конфигурация столбцов для сбора данных (номера столбцов, начиная с 0)
    config = {
        'Реестр': {
            'sheet_name': 'Лист1',
            'skip_rows': 2,
            'columns': {
                'number': 4,
                'date': 30
            }
        },
        'Реестр 2': {
            'sheet_name': 'Реестр 2',
            'skip_rows': 2,
            'columns': {
                'number': 3,
                'date1': 30,
                'date2': 50
            }
        },
        '!Трекер': {
            'sheet_name': content_customer,
            'skip_rows': 1,
            'columns': {
                'number': 0,
                'org': 46
            }
        }
    }

    data_p = read_sheet_data(
        content_folder_data / 'Реестр 1.xlsx',
        sheet_name=config['Реестр 1']['sheet_name'],
        skip_rows=config['Реестр 1']['skip_rows'],
        columns_config=config['Реестр 1']['columns'])

    data_2 = read_sheet_data(
        content_folder_data / 'Реестр 2.xlsx',
        sheet_name=config['Реестр 2']['sheet_name'],
        skip_rows=config['Реестр 2']['skip_rows'],
        columns_config=config['Реестр 2']['columns'])

    data_t = read_sheet_data(
        content_folder_data / '!Трекер.xlsx',
        sheet_name=config['!Трекер']['sheet_name'],
        skip_rows=config['!Трекер']['skip_rows'],
        columns_config=config['!Трекер']['columns'])

    # Сбор всех уникальных номеров
    all_numbers = set()
    for dataset in [data_p, data_2, data_t]:
        for row in dataset:
            num = row.get('number')
            if num is not None:
                all_numbers.add(num)

    sorted_numbers = sorted(all_numbers, key=lambda x: str(x))

    # Создание новой таблицы
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Выборка"

    # Заголовки
    headers = [
        'Номер',
        'Дата 1',
        'Дата 2',
        'Орг']
    ws_out.append(headers)

    # Форматирование заголовков
    bold_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws_out.cell(row=1, column=col).font = bold_font

    # Заполнение данных
    for number in sorted_numbers:
        date_p = None
        date_2 = None
        org = None

        # Поиск в Реестре 1
        for row in data_p:
            if row['number'] == number and row['date'] is not None:
                date_p = row['date']
                break

        # Поиск в Реестре 2
        for row in data_2:
            if row['number'] == number:
                # Берём первую непустую дату из двух столбцов
                date_2 = row.get('date1') or row.get('date2')
                if date_2 is not None:
                    break

        # Поиск в Трекере
        for row in data_t:
            if row['number'] == number and row.get('org') is not None:
                org = row['org']
                break

        ws_out.append([number, date_p, date_2, org])

    # Настройка ширины столбцов
    for col_letter, width in {'A': 7, 'B': 21, 'C': 28, 'D': 13}.items():
        ws_out.column_dimensions[col_letter].width = width

    # Форматирование дат
    date_format = 'dd.mm.yyyy'  # Короткий формат даты
    for row in ws_out.iter_rows(min_row=2, max_col=3, max_row=ws_out.max_row):
        for cell in row[1:3]:  # Столбцы B и C
            # Проверяем, что значение — это дата
            if isinstance(cell.value, datetime):
                cell.number_format = date_format

    # Сохранение файла, при необходимости
    # wb_out.save(folder / '!_Выборка из трекера и реестров 1 и 2.xlsx')

    # Проверка наличия лог-файла и его создание, если не существует
    log_path = dest_dir_1 / 'log.xlsx'
    if not log_path.exists():
        create_log_file(log_path)

    # Открытие лог-файла для добавления записей
    log_wb = load_workbook(log_path)

    # Чтение таблицы Excel, начиная со второй строки
    rows = list(wb_out['Выборка'].iter_rows(min_row=2, values_only=True))
    wb_out.close()

    for row in rows:
        current_id = row[0]  # Номер
        date_ready = row[1] if len(row) >= 2 else None  # Дата 1
        date_approval = row[2] if len(row) >= 3 else None  # Дата 2
        organization = row[3] if len(row) >= 4 else ''  # Орг

        # Проверка наличия обязательных полей
        if current_id is None or (not date_ready and not date_approval):
            continue

        # Подбираем нужную дату
        selected_date = date_ready if date_ready else date_approval
        # Пропускаем строку, если дата некорректная
        if not isinstance(selected_date, datetime):
            continue

        # Проверяем допустимые годы обработки
        year = selected_date.year
        if year not in [2024, 2025]:
            continue

        # Определение типа файла (1 или 2) и выбор директории его копирования
        is_1 = bool(date_ready)
        subfolder = '9. 1' if is_1 else '7. 2'
        destination_folder = dest_dir_1 if is_1 else dest_dir_2

        # Работа с файлом
        match year:
            case 2024:
                search_folder = doc24_dir / f"{int(current_id):04}" / subfolder
                found_pdfs = find_pdf_files(search_folder)
                if found_pdfs:
                    file_to_copy = get_most_recent_file(
                        found_pdfs, search_folder)
                    process_file(
                        file_to_copy,
                        search_folder,
                        current_id,
                        subfolder,
                        destination_folder,
                        log_wb,
                        organization,
                        dest_dir_1,
                        content_contractor_1,
                        content_contractor_2)
                else:
                    log_wb.active.append([
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'Не найден файл 2024.PDF в папке 2024',
                        current_id,
                        '',
                        subfolder.split(' ')[1]])

            case 2025:
                # 1. Сначала проверяем doc25_dir с точным совпадением номера
                search_folder = doc25_dir / f"{int(current_id):04}" / subfolder
                found_pdfs = find_pdf_files(search_folder)

                if not found_pdfs:
                    # 2. Если не найдено, проверяем doc25_dir с доп. символом
                    search_pattern = str(
                        doc25_dir / f"{int(current_id):04}*" / subfolder)
                    possible_folders = [
                        f for f in glob(search_pattern) if f != (
                            doc25_dir / f"{int(current_id):04}" / subfolder)]
                    if possible_folders:
                        # Берём первую подходящую папку
                        search_folder = possible_folders[0]
                        found_pdfs = find_pdf_files(search_folder)

                if not found_pdfs:
                    # 3. Если не найдено, проверяем doc24_dir с точным совп.
                    search_folder = doc24_dir / f"{
                        int(current_id):04}" / subfolder
                    found_pdfs = find_pdf_files(search_folder)

                if not found_pdfs:
                    # 4. Если не найдено, проверяем doc24_dir с доп. символом
                    search_pattern = str(
                        doc24_dir / f"{int(current_id):04}*" / subfolder)
                    possible_folders = [
                        f for f in glob(search_pattern) if f != (
                            doc24_dir / f"{int(current_id):04}" / subfolder)]
                    if possible_folders:
                        # Берём первую подходящую папку
                        search_folder = possible_folders[0]
                        found_pdfs = find_pdf_files(search_folder)

                # Обработка результата поиска
                if found_pdfs:
                    file_to_copy = get_most_recent_file(
                        found_pdfs, search_folder)
                    # Проверяем соответствие имени файла шаблону
                    if "2024" in str(search_folder) and not re.fullmatch(
                            r'\d{6}(2025|25).*\.pdf', file_to_copy):
                        log_wb.active.append([
                            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'Не найден файл 2025.PDF в папке 2024',
                            current_id,
                            '',
                            subfolder.split(' ')[1]])
                    else:
                        process_file(
                            file_to_copy,
                            search_folder,
                            current_id,
                            subfolder,
                            destination_folder,
                            log_wb,
                            organization,
                            dest_dir_1,
                            content_contractor_1,
                            content_contractor_2)
                else:
                    log_wb.active.append([
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'Не найден файл 2025.PDF в папке 2025',
                        current_id,
                        '',
                        subfolder.split(' ')[1]])

    # Добавляем строку в качестве разделителя и заканчиваем работу с лог-файлом
    log_wb.active.append(['работа завершена'])
    log_wb.save(log_path)
    log_wb.close()

    # Открываем директорию назначения
    open_file_and_folder(dest_dir_1 / 'log.xlsx')


# Функция для кнопки 5 - указание пути к лог-файлу по 1 и 2
def open_log_file_1(content_df_1):
    """
    На входе, функция получает:
    - content_df_1 - директория назначения для 1
    """
    # Формируем путь к лог-файлу
    content_df_1 = Path(content_df_1)
    log_file_path = (content_df_1 / 'log.xlsx')
    # Передаём путь к директории в функцию
    check_and_open_log(log_file_path)


# Вспомогательная функция для copy_ktd_button - считывание данных из файлов
def read_sheet_data(path, sheet_name, skip_rows=0, columns_config=None):
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet_name]

    row_iter = ws.iter_rows()

    # Пропускаем заданное количество строк
    for _ in range(skip_rows):
        next(row_iter)

    data = []
    for row_cells in row_iter:
        row_data = {}
        for key, col_idx in columns_config.items():
            if col_idx < len(row_cells):
                value = row_cells[col_idx].value
                # Преобразуем строки в даты, если это столбец с датой
                if key in (
                        'date', 'date1', 'date2') and isinstance(value, str):
                    try:
                        value = datetime.strptime(value, '%d.%m.%Y')
                    except ValueError:
                        value = None  # Если преобразование не удалось
                row_data[key] = value
            else:
                row_data[key] = None
        data.append(row_data)

    wb.close()
    return data


# Вспомогательная функция для copy_ktd_button - поиск PDF-файлов
def find_pdf_files(doc_dir):
    doc_dir_pathlib = Path(doc_dir)
    if doc_dir_pathlib.is_dir():
        return [
            file.name for file in doc_dir_pathlib.iterdir(
                ) if file.suffix.lower() == '.pdf']
    return []


# Вспомогательная функция для copy_ktd_button - выбор самого нового файла
def get_most_recent_file(file_list, search_folder):
    search_folder_pathlib = Path(search_folder)
    return max(
        file_list,
        key=lambda f: (search_folder_pathlib / f).stat().st_mtime if (
                        search_folder_pathlib / f).exists() else 0,
        default=None)


# Вспомогательная функция для copy_ktd_button - работа с PDF-файлом и лог
def process_file(
        file_to_copy,
        search_folder,
        current_id,
        subfolder,
        destination_folder,
        log_wb,
        org_dir,
        dest_dir_log,
        content_contractor_1,
        content_contractor_2):

    not_org_dir = destination_folder / 'Не распределены'
    not_org_file = not_org_dir / f"{int(current_id):04} - {file_to_copy}"

    # Если файл уже есть в "Не распределенных", но появилась новая информация
    if org_dir in (
            content_contractor_1,
            content_contractor_2) and not_org_file.exists():
        try:
            target_dir = destination_folder / org_dir
            target_dir.mkdir(parents=True, exist_ok=True)

            # Перемещаем файл в соответствующую папку
            shutil.move(
                not_org_file, target_dir / not_org_file.name)

            log_wb.active.append([
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                f"Перемещен из 'Не распределены' в '{org_dir}'",
                current_id,
                file_to_copy,
                subfolder.split(' ')[1],
                org_dir])

            return True
        except Exception as e:
            log_wb.active.append([
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                f'Ошибка перемещения: {str(e)}',
                current_id,
                file_to_copy])
            return False

    target_dir = Path(
        destination_folder / 'Не распределены' if org_dir not in (
            content_contractor_1, content_contractor_2) else org_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    src_file = Path(search_folder) / file_to_copy
    new_filename = f"{int(current_id):04} - {file_to_copy}"
    dest_file = target_dir / new_filename

    try:
        # Проверка лог-файла
        log_path = dest_dir_log / 'log.xlsx'
        logged_files = set()  # Множество для хранения уже скопированных файлов
        logged_ids = {}  # Словарь для хранения номеров и статусов
        if log_path.exists():
            wb = load_workbook(log_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Имя файла из четвёртого столбца
                logged_files.add(row[3])
                # Номер: (результат, имя файла)
                logged_ids[row[2]] = (row[1], row[3])
            wb.close()

        # Если файл уже скопирован, пропускаем
        if file_to_copy in logged_files:
            return True

        # Определяем сообщение для лога
        if current_id not in logged_ids:
            log_message = 'Файл скопирован'
        elif logged_ids[current_id][0] == 'Не найден файл .PDF':
            log_message = 'Файл скопирован'
        elif logged_ids[current_id][1]:  # Если ранее был скопирован файл
            log_message = 'Файл обновлён'
        else:
            log_message = 'Файл скопирован'

        # Копируем файл и вносим результат в лог-файл
        shutil.copy2(src_file, dest_file)
        log_wb.active.append([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            log_message,
            current_id,
            file_to_copy,
            subfolder.split(' ')[1],
            dest_file.parent.name])
        return True
    except Exception as e:
        log_wb.active.append([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            f'Ошибка копирования: {str(e)}',
            current_id,
            file_to_copy])
        return False


# Функция создания лог-файла (универсальная)
def create_log_file(log_path):
    wb = Workbook()
    ws = wb.active
    ws.append([
        'Дата',
        'Результат',
        'Номер',
        'Имя файла',
        'Тип документа',
        'Куда скопирован'])

    # Настройка стиля заголовка
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='thick'))

    # Установка ширины столбцов
    for col_letter, width in {
            'A': 19, 'B': 47, 'C': 6, 'D': 100, 'E': 15, 'F': 22}.items():
        ws.column_dimensions[col_letter].width = width

    # Сохранение файла
    wb.save(log_path)


# Функция для открытия лог-файла (универсальная)
def check_and_open_log(log_file_path):
    """
    На входе, функция получает:
    - log_file_path - имя файла, включая путь до него
    """
    # Проверяем, существует ли файл
    if log_file_path.is_file():
        # Открываем файл в зависимости от операционной системы
        system = platform.system()

        if system == 'Windows':
            subprocess.run(['explorer', log_file_path])  # nosec B603, B607
        elif system == 'Darwin':  # macOS
            subprocess.run(['open', log_file_path])  # nosec B603, B607
        else:  # Linux и другие системы
            subprocess.run(['xdg-open', log_file_path])  # nosec B603, B607
    else:
        messagebox.showwarning(
            'Предупреждение', f'Лог-файл не найден по пути: {log_file_path}')


# Функция для открытия файла и директории (универсальная)
def open_file_and_folder(output_filename):
    """
    На входе, функция получает:
    - output_filename - имя файла, включая путь до него
    """
    path = Path(output_filename).resolve()  # Полный абсолютный путь
    parent_dir = path.parent  # Получаем родительскую директорию

    # Открываем файл в зависимости от операционной системы
    system = platform.system()

    if system == 'Windows':
        # Открываем папку с выделенным файлом в проводнике
        subprocess.run(['explorer', '/select,', str(path)])  # nosec B603, B607
    elif system == 'Darwin':  # macOS
        # Показываем файл в Finder
        subprocess.run(['open', '-R', str(path)])  # nosec B603, B607
    else:  # Linux и другие системы
        # Показываем папку в файловом менеджере
        subprocess.run(['xdg-open', str(parent_dir)])  # nosec B603, B607
        # Открываем файл в ассоциированной программе
        subprocess.run(['xdg-open', str(path)])  # nosec B603, B607


# Функция для конвертации файла 2 в список
def transform_excel_list(input_fn, output_file):
    """
    На входе, функция получает:
    - input_fn - имя исходного файла, включая путь до него
    - output_file - имя выходного файла, включая путь до него
    """
    wb_out = None  # Инициализация переменной перед try-блоком

    # Проверяем наличие выходного файла и создаем его, если его нет
    try:
        wb_out = load_workbook(output_file)
    except FileNotFoundError:
        wb_out = Workbook()
        wb_out.save(output_file)
    finally:
        # Проверка, что переменная была инициализирована
        if wb_out is not None:
            wb_out.close()

    # Открываем исходную книгу
    wb_in = load_workbook(input_fn)
    ws_in = wb_in.active

    # Открываем новую книгу для записи
    wb_out = load_workbook(output_file)
    ws_out = wb_out.active

    # Определяем триггеры для обрезки
    trigger_start = '№ п/п'
    trigger_end = 'Исполнитель'

    # Инициализируем счётчики для индексов
    row_start = 1
    row_end = ws_in.max_row

    # Проходим по таблице и находим строки с триггерами
    for row_num in range(1, ws_in.max_row + 1):
        for col_num in range(1, ws_in.max_column + 1):
            cell_value = ws_in.cell(row=row_num, column=col_num).value

            if cell_value == trigger_start:
                # Начинаем копирование со строки с триггером
                row_start = row_num
            elif cell_value == trigger_end:
                # Заканчиваем копирование на две строки выше строки с триггером
                row_end = row_num - 2
                break  # Выходим из цикла, если нашли конец

    # Копируем строки между найденными границами
    for row_num in range(row_start, row_end + 1):
        for col_num in range(1, ws_in.max_column + 1):
            source_cell = ws_in.cell(row=row_num, column=col_num)
            dest_cell = ws_out.cell(
                row=row_num - row_start + 1, column=col_num)
            dest_cell.value = source_cell.value

            # Копируем стиль ячейки
            if source_cell.has_style:
                # Шрифт
                dest_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic)
                # Границы ячейки
                dest_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom)
                # Стиль содержимого ячейки
                dest_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text)

    # Сохраняем промежуточный файл
    wb_out.save(output_file)

    # Теперь открываем файл снова и удаляем пустые столбцы и строки
    wb_out = load_workbook(output_file)
    ws_out = wb_out.active

    # Перебираем все столбцы и проверяем, заполнены ли они
    empty_columns = []
    for col_idx in range(1, ws_out.max_column + 1):
        empty = True
        for row_idx in range(1, ws_out.max_row + 1):
            if ws_out.cell(row=row_idx, column=col_idx).value is not None:
                empty = False
                break
        if empty:
            empty_columns.append(col_idx)

    # Перебираем все строки и проверяем, заполнены ли они
    empty_rows = []
    for row_idx in range(1, ws_out.max_row + 1):
        empty = True
        for col_idx in range(1, ws_out.max_column + 1):
            if ws_out.cell(row=row_idx, column=col_idx).value is not None:
                empty = False
                break
        if empty:
            empty_rows.append(row_idx)

    # Удаляем пустые столбцы
    for col_idx in sorted(empty_columns, reverse=True):
        ws_out.delete_cols(col_idx, 1)

    # Удаляем пустые строки
    for row_idx in sorted(empty_rows, reverse=True):
        ws_out.delete_rows(row_idx, 1)

    # Задаём ширину определённых столбцов
    column_widths = {
        'A': 6,
        'B': 65,
        'C': 150,
        'D': 12,
        'E': 12}

    for col_name, width in column_widths.items():
        ws_out.column_dimensions[col_name].width = width

    # Задаём высоту строк в зависимости от наличия текста
    for row in ws_out.iter_rows():
        max_lines = 1  # Минимальное значение

        for cell in row:
            if cell.value and isinstance(cell.value, str):

                # Учитываем длину текста и ширину столбца
                col_width = ws_out.column_dimensions[
                    get_column_letter(cell.column)].width
                approx_chars_per_line = col_width / 2
                text_length = len(cell.value)
                length_based_lines = (
                    text_length / 1.3) / approx_chars_per_line
                # 2 - эмпирический коэффициент (примерное кол-во символов на
                # единицу ширины), можно регулировать, для более точного
                # подсчета символов

                # Берем максимальное значение
                current_lines = max(max_lines, length_based_lines)
                if current_lines > max_lines:
                    max_lines = current_lines

        # Устанавливаем высоту
        if max_lines > 1:
            # 13 единиц на строку - можно регулировать
            ws_out.row_dimensions[row[0].row].height = 13 * max_lines

    # Сохраняем промежуточный файл
    wb_out.save(output_file)

    # Загружаем книгу и чистим от лишних пробелов, табуляции и переносов строк
    workbook = load_workbook(output_file)
    sheet = workbook.active  # получаем активный лист

    # Проходим по всем ячейкам в указанном столбце
    for row in sheet.iter_rows(min_col=3, max_col=3):
        cell = row[0]  # В каждой строке берём только ячейку 3-го столбца
        if isinstance(cell.value, str):
            cell.value = re.sub(r'\s+', ' ', cell.value.strip())

    # Сохраняем изменения
    workbook.save(output_file)
    wb_in.close()
    wb_out.close()


# Формирование файла 2, на основе файла 1
def transform_excel_removal_equipment(input_fn, output_file, content_fwt_r):
    """
    На входе, функция получает:
    - input_fn - имя исходного файла, включая путь до него
    - output_file - имя выходного файла, включая путь до него
    - content_fwt_r - путь к информации
    """

    # Функция для изменения содержимого объединённой ячейки
    def set_merged_cell(cell_ref_content, value_content):
        cell_content = ws[cell_ref_content]
        cell_content.value = value_content

    # Настройка стилей первой группы ячеек
    def merge_and_style_1(ws_mas_1, merged_range_mas_1):
        # Настраиваем стили
        font = Font(name='Times New Roman', size=10, bold=True)
        alignment_mas_1 = Alignment(horizontal='center', vertical='center')

        # Объединяем ячейки
        ws_mas_1.merge_cells(merged_range_mas_1)

        # Получаем главную ячейку (первую в диапазоне)
        start_cell = merged_range_mas_1.split(':')[0]
        main_cell_1 = ws_mas_1[start_cell]

        # Применяем стили
        main_cell_1.font = font
        main_cell_1.alignment = alignment_mas_1

    # Настройка стилей второй группы ячеек
    def merge_and_style_2(ws_mas_2, merged_range_mas_2):
        # Настраиваем стили
        font = Font(name='Times New Roman', size=7)
        alignment_mas_2 = Alignment(horizontal='center', vertical='center')
        top_border = Border(top=Side(style='thin'))

        # Объединяем ячейки
        ws_mas_2.merge_cells(merged_range_mas_2)

        # Получаем главную ячейку (первую в диапазоне)
        start_cell = merged_range_mas_2.split(':')[0]
        main_cell_2 = ws_mas_2[start_cell]

        # Применяем стили
        main_cell_2.font = font
        main_cell_2.alignment = alignment_mas_2

        # Получаем границы диапазона
        min_col_mas_2, min_row_mas_2, max_col_mas_2, _ = range_boundaries(
            merged_range_mas_2)

        # Применяем отображение верхней границы
        for col in range(min_col_mas_2, max_col_mas_2 + 1):
            cell_mas_2 = ws.cell(row=min_row_mas_2, column=col)
            cell_mas_2.border = top_border

    # Обновляет объединённые ячейки
    def process_merged_ranges(
            ws_pmr, merged_ranges_pmr, value_pmr, alignment_pmr):
        for merged_range_pmr in merged_ranges_pmr:
            min_col, min_row, _, _ = range_boundaries(merged_range_pmr)
            main_cell = ws_pmr.cell(row=min_row, column=min_col)
            main_cell.value = value_pmr
            main_cell.alignment = alignment_pmr

    # 1) Загружаем файл
    wb = load_workbook(input_fn)
    sheet_name = 'Лист_1'
    ws = wb[sheet_name]

    # 2) Корректируем строку с первой позицией.
    # Настройки выравнивания и отображения текста
    alignment = Alignment(
        wrap_text=True,  # Включение переноса длинного текста
        horizontal='center',  # Включение выравнивания
        vertical='center')  # Включение выравнивания

    # Список объединённых диапазонов
    merged_ranges = [
        'D27:R27',
        'S27:Z27',
        'AA27:AE27']

    # Значение для всех объединённых ячеек
    target_value_null = '0'
    target_value_edit = ''

    # Название листа в Excel файле
    sheet_name = 'Выгрузка'

    # Индексы столбцов (начиная с 1)
    number_column_index = 2  # Первый столбец с числами
    data_column_index = 114  # Второй столбец с данными

    # Название, которое мы ищем в столбце с данными
    target_content = 'не требуется'

    # Загружаем Excel файл
    workbook = load_workbook(content_fwt_r)
    sheet = workbook[sheet_name]

    # Проходим по строкам в листе
    # Получаем имя файла из пути
    file_name = input_fn.stem.zfill(4)
    # Пропускаем заголовок
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Получаем число из первого столбца
        number = row[number_column_index - 1]
        # Получаем данные из второго столбца
        data = row[data_column_index - 1]

        # Проверяем, содержит ли строка искомое название
        if data is not None and target_content in str(data):
            # Преобразуем число в строку и дополняем нулями до 4 символов
            number_str = str(number).zfill(4) if number is not None else "0000"

            # Проверяем, начинаются ли первые 4 символа имени файла с числа
            if file_name == number_str:
                process_merged_ranges(
                    ws, merged_ranges,
                    target_value_null,
                    alignment)
                set_merged_cell(
                    'AF27', 'не предусмотрен')

                output_file = (
                    output_file.parent / "Для customer" / output_file.name)
        else:
            # Преобразуем число в строку и дополняем нулями до 4 символов
            number_str = str(number).zfill(4) if number is not None else "0000"

            # Проверяем, начинаются ли первые 4 символа имени файла с числа
            if file_name == number_str:
                process_merged_ranges(
                    ws, merged_ranges,
                    target_value_edit,
                    alignment)
                set_merged_cell('AF27', '')

    # Включение отображения длинного текста в ячейке AF27
    cell = ws['AF27']
    cell.alignment = alignment

    # 3) Определяем диапазон очистки: со строки 28 и до конца листа
    clear_min_row = 28
    clear_max_row = ws.max_row

    # Удаляем (разъединяем) объединённые ячейки в диапазоне
    merged_ranges = ws.merged_cells.ranges.copy()
    for merged_range in merged_ranges:
        first_row = merged_range.min_row
        last_row = merged_range.max_row
        if first_row >= clear_min_row and last_row <= clear_max_row:
            ws.unmerge_cells(str(merged_range))

    # Очищаем данные и форматирование в диапазоне
    default_font = Font(
        name='Times New Roman',
        size=11, bold=False,
        italic=False)
    default_fill = PatternFill(fill_type=None)
    default_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None))
    default_alignment = Alignment(
        horizontal='general',
        vertical='bottom')

    for row in ws.iter_rows(min_row=clear_min_row, max_row=clear_max_row):
        for cell in row:
            cell.value = None
            cell.font = default_font
            cell.fill = default_fill
            cell.border = default_border
            cell.alignment = default_alignment

    # Сбрасываем высоту строк, в которых была очистка данных
    for row in range(clear_min_row, clear_max_row + 1):
        if row in ws.row_dimensions:
            ws.row_dimensions[row].height = None

    # 4) Настраиваем и заполняем блок подписантов

    # Список диапазонов для объединения и стилизации первой группы ячеек
    merged_ranges_group_1 = [
        'B30:F30',
        'B35:F35',
        'G30:K30',
        'G35:K35',
        'X30:AD30',
        'X35:AD35',
        'AF30:AP30',
        'AF35:AP35']

    # Применяем функцию к каждому диапазону
    for merged_range_1 in merged_ranges_group_1:
        merge_and_style_1(ws, merged_range_1)

    # Список диапазонов для объединения и стилизации второй группы ячеек
    merged_ranges_group_2 = [
        'G31:K31',
        'G36:K36',
        'X31:AD31',
        'X36:AD36',
        'AF31:AP31',
        'AF36:AP36']

    # Применяем функцию к каждому диапазону
    for merged_range_2 in merged_ranges_group_2:
        merge_and_style_2(ws, merged_range_2)

        # Список ячеек и значений
        cells_values = [
            ('G22', 'ВЕДОМОСТЬ'),
            ('B30', 'Исполнитель'),
            ('G31', '(должность)'),
            ('X31', '(подпись)'),
            ('AF31', '(расшифровка подписи)'),
            ('B35', 'Заказчик'),
            ('G36', '(должность)'),
            ('X36', '(подпись)'),
            ('AF36', '(расшифровка подписи)')]

        # Вносим данные
        for cell_ref, value in cells_values:
            set_merged_cell(cell_ref, value)

    # 5) Устанавливаем высоту строк п.1 и блока подписантов
    row_heights = {
        27: 50,
        30: 33,
        35: 33}

    for row_num, height in row_heights.items():
        ws.row_dimensions[row_num].height = height

    # Сохраняем изменения в файл
    wb.save(output_file)


# ===== ===== ===== СЕРВЕР НАЧАЛО ===== ===== =====
# Пути к рабочим файлам
folder_with_tracker_1 = None
folder_with_tracker_2025 = None

# Целевая папка для копирования информации.
folder_data = None

# Исходная папка с файлами
source_folder_dp = None

# Исходные (оригинальные) имена файлов
original_file_name_tracker_2025 = 'Трекер.xlsb'
file_name_rep_t = 'Отчет.xlsx'
file_name_reg_2 = 'Реестр.xlsx'
file_name_reg_1 = 'Реестр.xlsx'

# Используемые наименования и названия
customer = 'customer'
contractor_1 = 'contractor_1'
contractor_2 = 'contractor_2'

# Приставки (суффиксы) для итоговых файлов
suffix_1 = '-1_customer'
suffix_2 = '-2_customer'

folder_for_customer = 'Для customer'

# Исходная папка с файлом
ip_folder = None

# Директория по умолчанию
default_folder_1 = None

# Исходные директории с документацией
root_folder_kd = None
root_folder_kd24 = None
root_folder_kd25 = None

# Целевые директории для скопированных файлов
destination_folder_1 = None
destination_folder_2 = None

# Путь к файлу
war_contractor_2 = None
# ===== ===== ===== СЕРВЕР КОНЕЦ ===== ===== =====

# Создание главного окна
root = tk.Tk()
root.title('The F.I.R. (File Interaction Runner)')

# Установка размеров окна
root.geometry('350x240')

# Кнопка Х - Копирование файлов
button_copy_file = tk.Button(
    root,
    text='Обновить файлы',
    command=lambda: copy_file_for_data(
        folder_data,
        ip_folder,
        source_folder_dp,
        root_folder_kd,
        original_file_name_tracker_2025,
        file_name_rep_t,
        file_name_reg_2,
        file_name_reg_1,
        customer))
button_copy_file.place(x=10, y=10, width=235, height=30)

# Кнопка XX - Открытие папки
button_open_data = tk.Button(
    root,
    text='Открыть Data',
    command=lambda: open_file_and_folder(folder_with_tracker_2025))
button_open_data.place(x=255, y=10, width=85, height=30)

# Текст над полем ввода
label1 = tk.Label(
    root,
    text='Работаем с папкой (изменить при необходимости):',
    font=('Arial', 10))
label1.place(x=10, y=45, width=330, height=20)

# Поле для ввода текста (Entry)
entry = tk.Entry(root)
entry.place(x=10, y=70, width=330, height=20)

# Установка значения по умолчанию
entry.insert(0, default_folder_1)  # Вставляем значение по умолчанию

# Кнопка 1 - Перемещение файла и создание списка
button1 = tk.Button(
    root,
    text='Обработать файлы в папке',
    command=lambda: copy_1_button(
        folder_with_tracker_1,
        folder_with_tracker_2025,
        war_contractor_2,
        customer,
        contractor_1,
        contractor_2,
        suffix_1,
        suffix_2,
        folder_for_customer))
button1.place(x=10, y=95, width=235, height=30)

# Кнопка 2 - Открытие лог-файла
button2 = tk.Button(
    root,
    text='Открыть лог-файл',
    command=open_log_file_)
button2.place(x=10, y=130, width=235, height=30)

# Кнопка 3 - Создание списка на основе выбранного файла
button3 = tk.Button(
    root,
    text='Сформировать из файла список',
    command=open_file_2)
button3.place(x=10, y=165, width=235, height=30)

# Надпись над текстовым полем для отображения обработанных файлов
processed_files_label = tk.Label(root, text='Обработаны:', font=('Arial', 10))
processed_files_label.place(x=255, y=90, width=85, height=20)

# Текстовое поле для отображения обработанных файлов
text_area = tk.Text(root)
text_area.place(x=255, y=110, width=85, height=85)

# Кнопка 4 - Поиск и копирование 1
button4 = tk.Button(
    root,
    text='Скопировать новые файлы',
    command=lambda: copy_ktd_button(
        root_folder_kd24,
        root_folder_kd25,
        destination_folder_1,
        destination_folder_2,
        folder_data,
        customer,
        contractor_1,
        contractor_2))
button4.place(x=10, y=200, width=235, height=30)

# Кнопка 5 - Открытие лога
button5 = tk.Button(
    root,
    text='Открыть лог',
    command=lambda: open_log_file_1(destination_folder_1))
button5.place(x=255, y=200, width=85, height=30)

# Запуск главного цикла обработки событий
root.mainloop()
