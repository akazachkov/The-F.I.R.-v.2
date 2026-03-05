# app/core/elements/excel_reader.py

import logging
from pathlib import Path
from typing import Union, List, Dict, Any, Optional, Callable

import openpyxl
from pyxlsb import open_workbook as open_xlsb

logger = logging.getLogger(__name__)


class ExcelReader:
    """
    Унифицированный класс для чтения данных из Excel файлов (xlsx, xlsb).
    Для xlsx используется режим read_only=True для обхода проблем с
    pivot-кэшами.
    """

    def __init__(self, file_path: Union[str, Path]):
        self.file_path = Path(file_path)
        self.ext = self.file_path.suffix.lower()
        self._workbook = None
        self._sheet = None
        self._is_xlsb = (self.ext == '.xlsb')

    def __enter__(self):
        self._open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._close()

    def _open(self):
        """Открывает книгу в зависимости от расширения."""
        if self._is_xlsb:
            self._workbook = open_xlsb(str(self.file_path))
        else:
            # Для xlsx и xlsm используем openpyxl в режиме read_only
            # Это позволяет избежать ошибок, связанных с pivot-кэшами
            self._workbook = openpyxl.load_workbook(
                self.file_path,
                data_only=True,
                read_only=True,
                keep_links=False  # не загружаем внешние связи
            )

    def _close(self):
        """Закрывает книгу."""
        if self._is_xlsb and self._workbook:
            self._workbook.close()
        elif self._workbook:
            self._workbook.close()
        self._workbook = None
        self._sheet = None

    def get_sheet_names(self) -> List[str]:
        if self._is_xlsb:
            return self._workbook.sheets
        else:
            return self._workbook.sheetnames

    def select_sheet(self, name: str):
        """Выбирает активный лист по имени."""
        if self._is_xlsb:
            self._sheet = self._workbook.get_sheet(name)
        else:
            self._sheet = self._workbook[name]
        return self

    def iter_rows(self, min_row: int = 1, values_only: bool = True):
        """
        Итератор по строкам текущего листа.
        Для xlsb возвращает списки значений (через .v).
        Для xlsx использует iter_rows().
        """
        if self._sheet is None:
            raise RuntimeError("Сначала вызовите select_sheet()")

        if self._is_xlsb:
            for row in self._sheet.rows():
                yield [cell.v for cell in row]
        else:
            # В режиме read_only доступен только iter_rows()
            for row in self._sheet.iter_rows(
                min_row=min_row, values_only=values_only
            ):
                yield row

    def get_mapping(
        self,
        sheet_name: str,
        key_col: int,
        value_col: int,
        start_row: int = 2,
        validator: Optional[Callable[[Any], bool]] = None,
        default: Any = None
    ) -> Dict[str, Any]:
        """
        Строит словарь из двух столбцов.

        :param sheet_name: имя листа
        :param key_col: номер столбца для ключа (1-based)
        :param value_col: номер столбца для значения (1-based)
        :param start_row: строка, с которой начинаются данные
        :param validator: функция, проверяющая, нужно ли включать значение
        :param default: значение по умолчанию, если валидатор не пройден
        :return: словарь {ключ: значение}
        """
        self.select_sheet(sheet_name)
        mapping = {}
        for row in self.iter_rows(min_row=start_row):
            if len(row) < max(key_col, value_col):
                continue
            key = row[key_col - 1]
            val = row[value_col - 1]
            if key is None:
                continue
            key_str = str(key).strip()
            if validator is not None:
                mapping[key_str] = val if validator(val) else default
            else:
                mapping[key_str] = val
        return mapping
