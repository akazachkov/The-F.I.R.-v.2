# app/core/elements/convert_register_to_list.py

import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Font, Alignment
from openpyxl.utils import get_column_letter


def transform_excel_list(input_fn, output_file):
    """
    Функция для конвертации ведомости в список.

    Args:
    :param input_fn (filename): имя исходного файла, включая путь до него
    :param output_file: имя выходного файла, включая путь до него
    """
    wb_out = None  # Инициализация переменной перед try-блоком

    # Проверяем наличие выходного файла и создаем его, если его нет
    try:
        wb_out = load_workbook(output_file)
    except FileNotFoundError:
        wb_out = Workbook()
        wb_out.save(output_file)
    finally:
        if wb_out is not None:
            wb_out.close()

    # Открываем исходную книгу
    wb_in = load_workbook(input_fn)
    ws_in = wb_in.active

    # Открываем новую книгу для записи
    wb_out = load_workbook(output_file)
    ws_out = wb_out.active

    # Определяем триггеры для обрезки
    trigger_start = '№ п/п'
    trigger_end = 'Исполнитель'

    # Инициализируем счетчики для индексов
    row_start = 0
    row_end = ws_in.max_row

    # Проходим по таблице и находим строки с триггерами
    for row_num in range(1, ws_in.max_row + 1):
        for col_num in range(1, ws_in.max_column + 1):
            cell_value = ws_in.cell(row=row_num, column=col_num).value

            if cell_value == trigger_start:
                # Начинаем со строки с триггером
                row_start = row_num
            elif cell_value == trigger_end:
                # Заканчиваем на две строки выше строки с триггером
                row_end = row_num - 2
                break

    # Копируем строки между найденными границами
    for row_num in range(row_start, row_end + 1):
        for col_num in range(1, ws_in.max_column + 1):
            source_cell = ws_in.cell(row=row_num, column=col_num)
            dest_cell = ws_out.cell(
                row=row_num - row_start + 1, column=col_num)
            dest_cell.value = source_cell.value

            # Копируем стиль ячейки
            if source_cell.has_style:
                dest_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic
                )  # Шрифт
                dest_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )  # Границы ячейки
                dest_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )  # Стиль содержимого ячейки

    # Сохраняем промежуточный файл
    wb_out.save(output_file)

    # Теперь открываем файл снова и удаляем пустые столбцы и строки
    wb_out = load_workbook(output_file)
    ws_out = wb_out.active

    # Перебираем все столбцы и проверяем, заполнены ли они
    empty_columns = []
    for col_idx in range(1, ws_out.max_column + 1):
        empty = True
        for row_idx in range(1, ws_out.max_row + 1):
            if ws_out.cell(row=row_idx, column=col_idx).value is not None:
                empty = False
                break
        if empty:
            empty_columns.append(col_idx)

    # Перебираем все строки и проверяем, заполнены ли они
    empty_rows = []
    for row_idx in range(1, ws_out.max_row + 1):
        empty = True
        for col_idx in range(1, ws_out.max_column + 1):
            if ws_out.cell(row=row_idx, column=col_idx).value is not None:
                empty = False
                break
        if empty:
            empty_rows.append(row_idx)

    # Удаляем пустые столбцы
    for col_idx in sorted(empty_columns, reverse=True):
        ws_out.delete_cols(col_idx, 1)

    # Удаляем пустые строки
    for row_idx in sorted(empty_rows, reverse=True):
        ws_out.delete_rows(row_idx, 1)

    # Задаем ширину определенных столбцов
    column_widths = {
        'A': 6,
        'B': 65,
        'C': 150,
        'D': 12,
        'E': 12
    }
    for col_name, width in column_widths.items():
        ws_out.column_dimensions[col_name].width = width

    # Задаем высоту строк в зависимости от наличия текста
    for row in ws_out.iter_rows():
        max_lines = 1  # Минимальное значение

        for cell in row:
            if cell.value and isinstance(cell.value, str):

                # Расчет учитывает длину текста и ширину столбца
                col_width = ws_out.column_dimensions[
                    get_column_letter(cell.column)].width
                approx_chars_per_line = col_width / 2  # Эмп. коэф.
                text_length = len(cell.value)
                length_based_lines = (
                    (text_length / 1.3) / approx_chars_per_line)
                # 2 - эмпирический коэффициент (примерное кол-во символов на
                # единицу ширины), можно регулировать

                # Берем максимальное значение
                current_lines = max(max_lines, length_based_lines)
                if current_lines > max_lines:
                    max_lines = current_lines

        # Устанавливаем высоту (13 единиц на строку, можно регулировать)
        if max_lines > 1:
            ws_out.row_dimensions[row[0].row].height = 13 * max_lines

    # Сохраняем промежуточный файл
    wb_out.save(output_file)

    # Чистим третий столбец от лишних пробелов, табуляции и переносов строк
    workbook = load_workbook(output_file)
    sheet = workbook.active  # Получаем активный лист (первый в файле)

    # Проходим по всем ячейкам в указанном столбце
    for row in sheet.iter_rows(min_col=3, max_col=3):
        cell = row[0]  # В каждой строке берем только ячейку 3-го столбца
        if isinstance(cell.value, str):
            cell.value = re.sub(r'\s+', ' ', cell.value.strip())

    # Сохраняем изменения
    workbook.save(output_file)
    wb_in.close()
    wb_out.close()
